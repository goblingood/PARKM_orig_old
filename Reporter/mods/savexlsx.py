# REPORTER create reports
# Copyright (C) 2013  AB

# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.

# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.

# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.


from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Alignment, Color, Fill
import datetime


def save_company_xlsx(file_name, data):
    wb = Workbook()
    ws = wb.get_active_sheet()
    column_width = []

    # cv - cell value
    # for j, cv in enumerate(columns):
    #     ws.cell(row = 2, column = j).value = cv
    #     ws.cell(row = 2, column = j).style.fill.fill_type = Fill.FILL_SOLID
    #     ws.cell(row = 2, column = j).style.fill.start_color.index = Color.BLUE

    for i, row in enumerate(data):
        for j, cv in enumerate(row):
            ws.cell(row=i, column=j).value = cv

            if len(column_width) > j:
                if len(str(cv)) > column_width[j]:
                    column_width[j] = len(str(cv))
            else:
                # column_width += [len(str(cv))]  # simpe => .append (see below)
                column_width.append(len(str(cv)))

            # set correct datatime _representation_ (not only data/number type/format) - there is some issue in openpyxl
            if isinstance(cv, datetime.datetime):
                ws.cell(row=i, column=j).style.number_format.format_code = 'dd-mm-yy hh:mm'
                ws.cell(row=i, column=j).style.alignment.horizontal = Alignment.HORIZONTAL_LEFT

            if j == 2:  # temorary fix cardid align AND ...E5 (...EX) see scientific notation
                ws.cell(row=i, column=j).set_value_explicit(cv)

    ws.cell(row=len(data)-1, column=6).style.number_format.format_code = '[hh]:mm:ss'  # temporary fix for Excel SUMM(timedelta)
    ws.cell(row=len(data)-1, column=6).style.fill.fill_type = Fill.FILL_SOLID
    ws.cell(row=len(data)-1, column=6).style.fill.start_color.index = Color.RED
    ws.cell(row=len(data)-1, column=7).style.fill.fill_type = Fill.FILL_SOLID
    ws.cell(row=len(data)-1, column=7).style.fill.start_color.index = Color.RED

    for i, width in enumerate(column_width):  # can be optimize [generator]
        ws.column_dimensions[get_column_letter(i+1)].width = width + 2  # 5 for best experience

    wb.save(file_name + '.xlsx')
